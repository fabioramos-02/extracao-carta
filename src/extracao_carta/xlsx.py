"""Manipulação de planilhas .xlsx preservando formatação."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .normalize import normalize


def enrich_column(
    in_path: Path,
    out_path: Path,
    mapping: dict[str, str],
    *,
    lookup_col: int = 2,
    write_col: int = 3,
    header: str = "categoria",
) -> tuple[int, int]:
    """Para cada aba, lê coluna `lookup_col`, normaliza e escreve `mapping[...]`
    em `write_col`. Linhas sem match recebem string vazia.
    Retorna (total_linhas_processadas, casadas).
    """
    wb = load_workbook(in_path)
    total = casados = 0
    for ws in wb.worksheets:
        ws.cell(row=1, column=write_col, value=header)
        for row in ws.iter_rows(min_row=2):
            if len(row) < lookup_col:
                continue
            val = row[lookup_col - 1].value
            if not val:
                continue
            total += 1
            slug = mapping.get(normalize(str(val)), "")
            if slug:
                casados += 1
            ws.cell(row=row[0].row, column=write_col, value=slug)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return total, casados
