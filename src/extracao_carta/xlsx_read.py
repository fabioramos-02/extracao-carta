"""Leitura de planilhas .xlsx (read-only)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def read_column(path: Path, col: int = 2, skip_header: bool = True) -> list[str]:
    """Lê valores não-vazios de uma coluna (1-based) em todas as abas."""
    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        start = 2 if skip_header else 1
        for row in ws.iter_rows(min_row=start, values_only=True):
            if len(row) < col:
                continue
            val = row[col - 1]
            if val:
                out.append(str(val))
    return out
