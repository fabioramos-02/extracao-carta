"""CLI: gera HTML estático do BI a partir do XLSX de cartas com PDF/vídeo.

CSS do Design System é inlinado no HTML — arquivo resultante é autocontido.

Uso (sem argumentos usa caminhos padrão da demanda 006):
    python scripts/generate_bi.py
    python scripts/generate_bi.py \\
        --ds-css node_modules/@design-system-ms/ds-sis/dist/css/ds-sis.css
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date
from pathlib import Path

from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
_DEMANDA = ROOT / "demandas/2026-06/006-cartas-pdf-video"

_DEFAULT_IN  = _DEMANDA / "output/cartas-pdf-video.xlsx"
_DEFAULT_OUT = _DEMANDA / "bi/index.html"
_DEFAULT_TPL = _DEMANDA / "bi/template.html"

_COLS = (
    "sigla_orgao", "titulo_servico", "categoria", "url_carta",
    "secao", "tipo", "url_midia", "logo_politica",
)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Gera BI HTML estático a partir do XLSX.")
    p.add_argument("--in",       dest="in_path",  type=Path, default=_DEFAULT_IN)
    p.add_argument("--out",      dest="out_path", type=Path, default=_DEFAULT_OUT)
    p.add_argument("--template", type=Path,        default=_DEFAULT_TPL)
    p.add_argument("--ds-css",   dest="ds_css",   type=Path, default=None,
                   help="ds-sis.css inlinado no HTML. Path: dist/css/ds-sis.css")
    return p.parse_args()


def _read_rows(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(dict(zip(_COLS, (c or "" for c in row))))
    return rows


def _group_by_carta(rows: list[dict]) -> list[dict]:
    """Agrupa linhas por carta — uma entrada por serviço com lista de links."""
    groups: dict[str, dict] = defaultdict(lambda: {"links": [], "logo_politica": ""})
    order: list[str] = []
    for r in rows:
        key = r["titulo_servico"]
        if key not in groups or "sigla_orgao" not in groups[key]:
            groups[key].update({
                "sigla_orgao":    r["sigla_orgao"],
                "titulo_servico": r["titulo_servico"],
                "categoria":      r["categoria"],
                "url_carta":      r["url_carta"],
                "logo_politica":  r.get("logo_politica", ""),
            })
        if key not in order:
            order.append(key)
        groups[key]["links"].append({
            "secao":    r["secao"],
            "tipo":     r["tipo"],
            "url_midia": r["url_midia"],
        })
    for key in order:
        tipos = sorted({lk["tipo"] for lk in groups[key]["links"]})
        groups[key]["tipos_str"] = " ".join(tipos)
    return [groups[k] for k in order]


def _stats(rows: list[dict], cartas: list[dict]) -> dict:
    return {
        "total_cartas": len(cartas),
        "total_pdf":    sum(1 for r in rows if r["tipo"] == "pdf"),
        "total_video":  sum(1 for r in rows if r["tipo"] == "video"),
        "gerado_em":    date.today().strftime("%d/%m/%Y"),
    }


def main() -> int:
    args = parse_args()

    if not args.in_path.exists():
        print(f"[erro] XLSX não encontrado: {args.in_path}")
        print("       Rode primeiro: python scripts/extract_cartas_pdf_video.py")
        return 1

    rows = _read_rows(args.in_path)
    cartas = _group_by_carta(rows)

    css_content = ""
    if args.ds_css and Path(args.ds_css).exists():
        css_content = Path(args.ds_css).read_text(encoding="utf-8")
    elif args.ds_css:
        print(f"[warn] ds-css não encontrado: {args.ds_css} — HTML sem DS")

    args.out_path.parent.mkdir(parents=True, exist_ok=True)
    env = Environment(
        loader=FileSystemLoader(str(args.template.parent)),
        autoescape=True,
    )
    html = env.get_template(args.template.name).render(
        cartas=cartas,
        stats=_stats(rows, cartas),
        css_content=css_content,
    )
    args.out_path.write_text(html, encoding="utf-8")

    s = _stats(rows, cartas)
    print(f"[ok] BI gerado: {args.out_path}")
    print(f"[ok] {s['total_cartas']} cartas | {s['total_pdf']} PDFs | {s['total_video']} vídeos")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
